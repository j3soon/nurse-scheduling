"""Tests for export formatting rules in XLSX export."""

# This file is part of Nurse Scheduling Project, see <https://github.com/j3soon/nurse-scheduling>.
#
# Copyright (C) 2023-2026 Johnson Sun
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

# This test is mostly AI generated.

import os
import sys
from io import BytesIO

from openpyxl import load_workbook

# Add the project root to the Python path so imports will work when running directly
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from nurse_scheduling import exporter, schedule


def test_export_formatting_rules_apply_to_rows_columns_headers_and_cells():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-03
people:
  items:
    - id: n1
    - id: n2
    - id: n3
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
    qualifiedPeople: ALL
    date: ALL
    weight: -1
  - type: shift request
    person: n3
    date: ["2025-01-01", "2025-01-02", "2025-01-03"]
    shiftType: D
    weight: 100
export:
  formatting:
    - type: row header
      targets: [n2]
      backgroundColor: "#f97316"
    - type: row
      targets: [n2]
      backgroundColor: "#06b6d4"
      bottomBorderColor: "#ef4444"
    - type: cell
      targets: [D]
      backgroundColor: "#1f2937"
    - type: column header
      targets: ["2025-01-01"]
      backgroundColor: "#a855f7"
    - type: column
      targets: ["2025-01-02"]
      backgroundColor: "#84cc16"
      bottomBorderColor: "#3b82f6"
"""

    df, _solution, _score, _status, cell_export_info = schedule(yaml_content, prettify=False)
    output = BytesIO()
    exporter.export_to_excel(df, output, cell_export_info)

    if os.getenv("WRITE_XLSX_ARTIFACT") == "1":
        artifact_path = os.path.join(
            os.path.dirname(__file__),
            "artifacts",
            "test_export_formatting.xlsx",
        )
        os.makedirs(os.path.dirname(artifact_path), exist_ok=True)
        with open(artifact_path, "wb") as f:
            f.write(output.getvalue())
        print(f"Wrote XLSX artifact: {artifact_path}")

    wb = load_workbook(output)
    ws = wb.active

    # Row target is n2 (Excel row 4). Row rule is after row-header rule, so it wins.
    assert ws["A4"].fill.fgColor.rgb == "FF06B6D4"
    assert ws["B4"].fill.fgColor.rgb == "FF06B6D4"
    assert ws["D4"].fill.fgColor.rgb == "FF06B6D4"

    # Column-header target is date 2025-01-01 (Excel column B), first row only.
    assert ws["B1"].fill.fgColor.rgb == "FFA855F7"

    # Column target is 2025-01-02 (Excel column C), and should style the entire column.
    assert ws["C1"].fill.fgColor.rgb == "FF84CC16"
    assert ws["C2"].fill.fgColor.rgb == "FF84CC16"
    assert ws["C3"].fill.fgColor.rgb == "FF84CC16"
    assert ws["C4"].fill.fgColor.rgb == "FF84CC16"
    assert ws["C5"].fill.fgColor.rgb == "FF84CC16"
    # Score/Status summary rows should not be affected by full-column styling.
    assert ws["C6"].fill.fgColor.rgb == "00000000"
    assert ws["C7"].fill.fgColor.rgb == "00000000"

    # Cell rule for D applies to assigned schedule cells that are not overridden by column style.
    assert ws["B5"].fill.fgColor.rgb == "FF1F2937"
    assert ws["D5"].fill.fgColor.rgb == "FF1F2937"

    # A4 bottom border from row rule.
    assert ws["A4"].border.bottom.color is not None
    assert ws["A4"].border.bottom.color.rgb == "FFEF4444"
    # C4 bottom border from column rule overriding row rule.
    assert ws["C4"].border.bottom.color is not None
    assert ws["C4"].border.bottom.color.rgb == "FF3B82F6"
