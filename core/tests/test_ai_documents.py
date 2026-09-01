"""Tests for bounded AI document extraction."""

# This file is part of Nurse Scheduling Project, see <https://github.com/j3soon/nurse-scheduling>.
#
# Copyright (C) 2023-2026 Johnson Sun
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

# This test is mostly AI generated.

from collections.abc import AsyncIterator, Sequence
from dataclasses import replace
from datetime import date
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from pypdf import PdfWriter

from nurse_scheduling.ai.app import create_app
from nurse_scheduling.ai.config import AiSettings
from nurse_scheduling.ai.documents import (
    MAX_XLSX_ARCHIVE_ENTRIES,
    DocumentExtractionLimits,
    DocumentLimitError,
    InvalidDocumentError,
    extract_document_text,
)
from nurse_scheduling.ai.provider import ChatMessage, TextDelta
from nurse_scheduling.ai.sandbox.fake import FakeSandboxFactory


class FakeProvider:
    """Record the extracted provider prompt."""

    def __init__(self) -> None:
        self.calls: list[list[ChatMessage]] = []

    async def stream_events(self, messages: Sequence[ChatMessage], tools=None) -> AsyncIterator[TextDelta]:
        self.calls.append(list(messages))
        yield TextDelta("Document answer")


def _limits(**overrides: int) -> DocumentExtractionLimits:
    limits = DocumentExtractionLimits(
        max_text_chars=50_000,
        max_pdf_pages=100,
        max_xlsx_sheets=20,
        max_xlsx_cells=100_000,
        max_xlsx_uncompressed_bytes=50_000_000,
    )
    return replace(limits, **overrides)


def _pdf_with_pages(page_texts: list[str]) -> bytes:
    """Build a small dependency-free PDF fixture with WinAnsi text pages."""
    font_object = 3 + len(page_texts) * 2
    page_objects = [3 + index * 2 for index in range(len(page_texts))]
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        (
            b"<< /Type /Pages /Kids ["
            + b" ".join(f"{page_object} 0 R".encode() for page_object in page_objects)
            + f"] /Count {len(page_texts)} >>".encode()
        ),
    ]
    for page_object, text in zip(page_objects, page_texts, strict=True):
        content_object = page_object + 1
        escaped_text = text.encode("cp1252").replace(b"\\", b"\\\\").replace(b"(", b"\\(").replace(b")", b"\\)")
        content = b"BT /F1 12 Tf 72 720 Td (" + escaped_text + b") Tj ET"
        objects.extend(
            [
                (
                    b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
                    + f"/Resources << /Font << /F1 {font_object} 0 R >> >> ".encode()
                    + f"/Contents {content_object} 0 R >>".encode()
                ),
                b"<< /Length " + str(len(content)).encode() + b" >>\nstream\n" + content + b"\nendstream",
            ]
        )
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>")
    result = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for index, value in enumerate(objects, start=1):
        offsets.append(len(result))
        result.extend(f"{index} 0 obj\n".encode())
        result.extend(value)
        result.extend(b"\nendobj\n")
    xref_offset = len(result)
    result.extend(f"xref\n0 {len(objects) + 1}\n".encode())
    result.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        result.extend(f"{offset:010d} 00000 n \n".encode())
    result.extend(f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode())
    return bytes(result)


def _pdf_with_text(text: str) -> bytes:
    return _pdf_with_pages([text])


def _rewrite_zip_entry(data: bytes, filename: str, replacement: bytes) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(data)) as source, ZipFile(output, "w", ZIP_DEFLATED) as target:
        for entry in source.infolist():
            content = replacement if entry.filename == filename else source.read(entry.filename)
            target.writestr(entry, content)
    return output.getvalue()


def _xlsx_with_formula() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Coverage"
    worksheet["A1"] = "Day"
    worksheet["A2"] = 1
    worksheet["B2"] = 2
    worksheet["C2"] = "=SUM(A2:B2)"
    worksheet["C3"] = "=A2"
    initial = BytesIO()
    workbook.save(initial)

    with ZipFile(BytesIO(initial.getvalue())) as archive:
        worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
    original = b'<c r="C2"><f>SUM(A2:B2)</f><v /></c>'
    assert original in worksheet_xml
    worksheet_xml = worksheet_xml.replace(original, b'<c r="C2"><f>SUM(A2:B2)</f><v>3</v></c>')
    return _rewrite_zip_entry(initial.getvalue(), "xl/worksheets/sheet1.xml", worksheet_xml)


def _xlsx_with_cell_variants() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Résumé"
    worksheet["A1"] = "你好"
    worksheet["A2"] = date(2026, 8, 26)
    worksheet["A3"] = True
    worksheet["A4"] = "#DIV/0!"
    worksheet["A4"].data_type = "e"
    worksheet.merge_cells("B1:C1")
    worksheet["B1"] = "Merged"
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "Hidden value"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _xlsx_with_formula_variants() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Formulas"
    worksheet["A2"] = 1
    worksheet["B2"] = 2
    worksheet["A3"] = 3
    worksheet["B3"] = 4
    worksheet["C2"] = "=A2+B2"
    worksheet["C3"] = "=A3+B3"
    worksheet["D2"] = "=A2:A3*2"
    worksheet["D3"] = 6
    worksheet["E2"] = "='[external.xlsx]Sheet1'!A1"
    initial = BytesIO()
    workbook.save(initial)
    with ZipFile(BytesIO(initial.getvalue())) as archive:
        worksheet_xml = archive.read("xl/worksheets/sheet1.xml")

    replacements = {
        b'<c r="C2"><f>A2+B2</f><v /></c>': b'<c r="C2"><f t="shared" ref="C2:C3" si="0">A2+B2</f><v>3</v></c>',
        b'<c r="C3"><f>A3+B3</f><v /></c>': b'<c r="C3"><f t="shared" si="0"></f><v>7</v></c>',
        b'<c r="D2"><f>A2:A3*2</f><v /></c>': b'<c r="D2"><f t="array" ref="D2:D3">A2:A3*2</f><v>2</v></c>',
    }
    for original, replacement in replacements.items():
        assert original in worksheet_xml
        worksheet_xml = worksheet_xml.replace(original, replacement)
    return _rewrite_zip_entry(initial.getvalue(), "xl/worksheets/sheet1.xml", worksheet_xml)


def _xlsx_with_encrypted_entry_flag() -> bytes:
    data = bytearray(_xlsx_with_formula())
    offset = 0
    while (offset := data.find(b"PK\x03\x04", offset)) >= 0:
        flags = int.from_bytes(data[offset + 6 : offset + 8], "little") | 0x1
        data[offset + 6 : offset + 8] = flags.to_bytes(2, "little")
        offset += 4
    offset = 0
    while (offset := data.find(b"PK\x01\x02", offset)) >= 0:
        flags = int.from_bytes(data[offset + 8 : offset + 10], "little") | 0x1
        data[offset + 8 : offset + 10] = flags.to_bytes(2, "little")
        offset += 4
    return bytes(data)


def _xlsx_with_too_many_entries() -> bytes:
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for index in range(MAX_XLSX_ARCHIVE_ENTRIES + 1):
            archive.writestr(f"entry-{index}", b"")
    return output.getvalue()


def _pdf_from_writer(*, pages: int = 1, password: str | None = None) -> bytes:
    writer = PdfWriter()
    for _ in range(pages):
        writer.add_blank_page(width=612, height=792)
    if password is not None:
        writer.encrypt(password)
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def test_pdf_and_xlsx_are_extracted_for_one_provider_request() -> None:
    provider = FakeProvider()
    settings = AiSettings(
        provider_base_url="https://provider.example/v1",
        provider_api_key="test-token",
        provider_model="test-model",
        cookie_secure=False,
    )
    client = TestClient(create_app(settings=settings, provider=provider, sandbox_factory=FakeSandboxFactory()))
    session = client.post("/sessions", json={"schedule_yaml": "description: test"})

    response = client.post(
        f"/sessions/{session.json()['id']}/messages",
        data={"message": "Read both documents."},
        files=[
            ("documents", ("notes.pdf", _pdf_with_text("Alice works Monday"), "application/pdf")),
            (
                "documents",
                (
                    "coverage.xlsx",
                    _xlsx_with_formula(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            ),
        ],
    )

    assert response.status_code == 200
    prompt = provider.calls[0][-1]["content"]
    assert isinstance(prompt, str)
    assert "Alice works Monday" in prompt
    assert 'C2: formula \\"=SUM(A2:B2)\\", cached value 3' in prompt
    assert 'C3: formula \\"=A2\\", cached value unavailable' in prompt


def test_pdf_extraction_marks_pages_without_embedded_text() -> None:
    text = extract_document_text("scan.pdf", _pdf_from_writer(), _limits())

    assert text == "--- PDF page 1 ---\n[No extractable text. OCR was not performed.]"


def test_pdf_extraction_preserves_page_order_and_non_ascii_text() -> None:
    text = extract_document_text("notes.pdf", _pdf_with_pages(["First café", "Second Åsa"]), _limits())

    assert text == "--- PDF page 1 ---\nFirst café\n--- PDF page 2 ---\nSecond Åsa"


@pytest.mark.parametrize(
    ("data", "limits", "expected_error"),
    [
        (_pdf_from_writer(pages=2), _limits(max_pdf_pages=1), "PDF attachment has too many pages."),
        (_pdf_from_writer(password="secret"), _limits(), "Encrypted PDF attachments are not supported."),
        (b"not pdf", _limits(), "PDF content does not match its filename."),
        (b"%PDF-1.7\nbroken", _limits(), "PDF attachment is invalid or unsupported."),
    ],
    ids=["too-many-pages", "encrypted", "not-a-pdf", "malformed"],
)
def test_pdf_rejects_unsupported_or_bounded_content(
    data: bytes,
    limits: DocumentExtractionLimits,
    expected_error: str,
) -> None:
    error_type = DocumentLimitError if "too many" in expected_error else InvalidDocumentError
    with pytest.raises(error_type, match=expected_error):
        extract_document_text("notes.pdf", data, limits)


def test_xlsx_extraction_preserves_formulas_and_cached_values() -> None:
    text = extract_document_text("coverage.xlsx", _xlsx_with_formula(), _limits())

    assert '--- XLSX sheet "Coverage" ---' in text
    assert 'A1: "Day"' in text
    assert 'C2: formula "=SUM(A2:B2)", cached value 3' in text
    assert 'C3: formula "=A2", cached value unavailable' in text


def test_xlsx_extraction_includes_cell_types_merged_cells_and_hidden_sheets() -> None:
    text = extract_document_text("variants.xlsx", _xlsx_with_cell_variants(), _limits())

    assert '--- XLSX sheet "Résumé" ---' in text
    assert 'A1: "你好"' in text
    assert 'A2: "2026-08-26T00:00:00"' in text
    assert "A3: true" in text
    assert 'A4: "#DIV/0!"' in text
    assert 'B1: "Merged"' in text
    assert "C1:" not in text
    assert '--- XLSX sheet "Hidden" ---' in text
    assert 'A1: "Hidden value"' in text


def test_xlsx_extraction_preserves_shared_array_and_external_reference_formulas() -> None:
    text = extract_document_text("formulas.xlsx", _xlsx_with_formula_variants(), _limits())

    assert 'C2: formula "=A2+B2", cached value 3' in text
    assert 'C3: formula "=A3+B3", cached value 7' in text
    assert 'D2: formula "=A2:A3*2", cached value 2' in text
    assert "E2: formula \"='[external.xlsx]Sheet1'!A1\", cached value unavailable" in text


@pytest.mark.parametrize(
    ("limits", "expected_error"),
    [
        (_limits(max_xlsx_sheets=0), "XLSX attachment has too many sheets."),
        (_limits(max_xlsx_cells=2), "XLSX attachment contains too many cells."),
        (_limits(max_xlsx_uncompressed_bytes=1), "XLSX attachment expands beyond the allowed size."),
    ],
)
def test_xlsx_enforces_structural_limits(limits: DocumentExtractionLimits, expected_error: str) -> None:
    with pytest.raises(DocumentLimitError, match=expected_error):
        extract_document_text("coverage.xlsx", _xlsx_with_formula(), limits)


def test_xlsx_rejects_content_that_is_not_an_ooxml_workbook() -> None:
    with pytest.raises(InvalidDocumentError, match="XLSX content does not match its filename"):
        extract_document_text("coverage.xlsx", b"not xlsx", _limits())


@pytest.mark.parametrize(
    ("data", "expected_error"),
    [
        (_xlsx_with_too_many_entries(), "XLSX attachment contains too many archive entries."),
        (_xlsx_with_encrypted_entry_flag(), "Encrypted XLSX attachments are not supported."),
        (
            _rewrite_zip_entry(
                _xlsx_with_formula(),
                "xl/worksheets/sheet1.xml",
                b"<worksheet><broken>",
            ),
            "XLSX attachment is invalid or unsupported.",
        ),
    ],
    ids=["too-many-entries", "encrypted-entry", "malformed-sheet"],
)
def test_xlsx_rejects_unsafe_or_malformed_archives(data: bytes, expected_error: str) -> None:
    with pytest.raises((DocumentLimitError, InvalidDocumentError), match=expected_error):
        extract_document_text("coverage.xlsx", data, _limits())


@pytest.mark.parametrize(
    ("filename", "data"),
    [
        ("notes.pdf", _pdf_with_text("Text beyond the extraction limit")),
        ("coverage.xlsx", _xlsx_with_formula()),
    ],
    ids=["pdf", "xlsx"],
)
def test_binary_document_extraction_enforces_prompt_size_limit(filename: str, data: bytes) -> None:
    with pytest.raises(DocumentLimitError, match="Document extracted text is too large"):
        extract_document_text(filename, data, _limits(max_text_chars=10))


def test_text_extraction_has_a_separate_prompt_size_limit() -> None:
    with pytest.raises(DocumentLimitError, match="Document extracted text is too large"):
        extract_document_text("notes.txt", b"12345", _limits(max_text_chars=4))
