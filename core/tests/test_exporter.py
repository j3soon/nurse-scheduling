"""Focused tests for exporter helpers and formatting edge cases."""

# This file is part of Nurse Scheduling Project, see <https://github.com/j3soon/nurse-scheduling>.
#
# Copyright (C) 2023-2026 Johnson Sun
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

# This test is mostly AI generated.

import os
import sys
from io import BytesIO
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook

# Add the project root to the Python path so imports will work when running directly
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from nurse_scheduling import exporter, schedule


def test_export_to_csv_writes_utf8_bom():
    df = pd.DataFrame([["A", "B"], ["C", "D"]])
    output = BytesIO()

    exporter.export_to_csv(df, output)

    payload = output.getvalue()
    assert payload.startswith(b"\xef\xbb\xbf")
    assert payload.decode("utf-8-sig") == "A,B\nC,D\n"


def test_export_to_excel_supports_legacy_comment_info_shape():
    df = pd.DataFrame([["x"]])
    output = BytesIO()

    # Legacy shape: {(row, col): [weights]}
    exporter.export_to_excel(df, output, {(1, 1): [3, 7]})

    wb = load_workbook(output)
    ws = wb.active
    assert ws["A1"].comment is not None
    assert "Weights of unmet single-style requests: 10" in ws["A1"].comment.text
    assert "3, 7" in ws["A1"].comment.text


def test_export_to_excel_applies_style_and_font_contrast():
    df = pd.DataFrame([["dark", "light"]])
    output = BytesIO()
    cell_export_info = {
        "comments": {},
        "styles": {
            (1, 1): {"backgroundColor": "#111111"},
            (1, 2): {
                "backgroundColor": "#f5f5f5",
                "bottomBorderColor": "#0ea5e9",
                "rightBorderColor": "#9ca3af",
            },
        },
    }

    exporter.export_to_excel(df, output, cell_export_info)

    wb = load_workbook(output)
    ws = wb.active
    assert ws["A1"].fill.fgColor.rgb == "FF111111"
    assert ws["A1"].font.color is not None
    assert ws["A1"].font.color.rgb == "FFFFFFFF"
    assert ws["B1"].fill.fgColor.rgb == "FFF5F5F5"
    assert ws["B1"].font.color is not None
    assert ws["B1"].font.color.rgb == "FF000000"
    assert ws["B1"].border.bottom.color is not None
    assert ws["B1"].border.bottom.color.rgb == "FF0EA5E9"
    assert ws["B1"].border.bottom.style == "medium"
    assert ws["B1"].border.right.color is not None
    assert ws["B1"].border.right.color.rgb == "FF9CA3AF"
    assert ws["B1"].border.right.style == "medium"


def test_prettify_generates_comment_info_for_unmet_single_style_requests():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-02
people:
  items:
    - id: n1
      history: [D]
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
    qualifiedPeople: ALL
    date: ALL
    weight: -1
  - type: shift request
    person: n1
    date: ["2025-01-01"]
    shiftType: D
    weight: -10
"""

    df, _solution, _score, _status, cell_export_info = schedule(yaml_content, prettify=True)
    assert hasattr(df, "to_excel")
    assert "comments" in cell_export_info
    assert "styles" in cell_export_info
    assert cell_export_info["comments"]

    first_target_cell = next(iter(cell_export_info["comments"].keys()))
    output = BytesIO()
    exporter.export_to_excel(df, output, cell_export_info)

    wb = load_workbook(output)
    ws = wb.active
    row, col = first_target_cell
    comment = ws.cell(row=row, column=col).comment
    assert comment is not None
    assert "Weight of unmet single-style request: 10" in comment.text


def test_invalid_row_target_in_export_formatting_raises():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
    qualifiedPeople: ALL
    date: ALL
    weight: -1
export:
  formatting:
    - type: row
      targets: [unknown_person]
      backgroundColor: "#22c55e"
"""

    with pytest.raises(ValueError, match="Invalid person identifier"):
        schedule(yaml_content, prettify=False)


def test_invalid_cell_target_in_export_formatting_raises():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
    qualifiedPeople: ALL
    date: ALL
    weight: -1
export:
  formatting:
    - type: cell
      targets: [UNKNOWN_SHIFT]
      backgroundColor: "#ef4444"
"""

    with pytest.raises(ValueError, match="Invalid shift type identifier"):
        schedule(yaml_content, prettify=False)


def test_date_headers_format_with_year_boundary():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2024-12-31
    endDate: 2025-01-02
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
"""

    df, _solution, _score, _status, _cell_export_info = schedule(yaml_content, prettify=False)
    assert df.iloc[0, 1] == "2024/12/31"
    assert df.iloc[0, 2] == "2025/1/1"
    assert df.iloc[0, 3] == "2025/1/2"


def test_prettify_off_annotations_and_workday_freeday_headers():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-03
  groups:
    - id: WORKDAY
      members: [2025-01-02, 2025-01-03]
    - id: FREEDAY
      members: [2025-01-01]
people:
  items:
    - id: n1
      history: [D]
    - id: n2
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
  - type: shift request
    person: n1
    date: ["2025-01-01"]
    shiftType: [D]
    weight: 0
  - type: shift request
    person: n1
    date: ["2025-01-01"]
    shiftType: [OFF]
    weight: -5
export:
  extraColumns:
    - type: count
      header: OFF (WORKDAY)
      countShiftTypes: [OFF]
      countDates: [WORKDAY]
    - type: count
      header: OFF (FREEDAY)
      countShiftTypes: [OFF]
      countDates: [FREEDAY]
  extraRows:
    - type: count
      header: OFF Count
      countShiftTypes: [OFF]
      countPeople: [ALL]
"""
    styled_df, _solution, _score, _status, _cell_export_info = schedule(yaml_content, prettify=True)
    df = styled_df.data

    # Weight-0 shift request should be ignored by prettify markers,
    # while OFF request should still annotate the cell.
    target_cell = str(df.iloc[2, 2])
    assert "[OFF]" in target_cell
    assert "[D]" not in target_cell

    # History fallback branch for person without history.
    assert df.iloc[3, 1] == ""

    # Workday/freeday summary headers should be present when both groups are found.
    headers = list(df.iloc[1, :])
    assert "OFF (WORKDAY)" in headers
    assert "OFF (FREEDAY)" in headers
    assert df.iloc[7, 0] == "OFF Count"
    assert df.iloc[7, 2] == 2


def test_build_custom_export_style_info_ignores_out_of_bounds_targets():
    ctx = SimpleNamespace(
        export=SimpleNamespace(
            formatting=[
                SimpleNamespace(
                    type="row",
                    targets=["n1"],
                    backgroundColor="#22c55e",
                    bottomBorderColor=None,
                    rightBorderColor=None,
                )
            ]
        ),
        map_pid_p={"n1": [0]},
        map_did_d={},
        map_sid_s={},
    )

    # n_rows=0 forces set_style to hit out-of-bounds guard and skip writes.
    style_map = exporter._build_custom_export_style_info(
        ctx,
        n_rows=0,
        n_cols=1,
        n_leading_rows=2,
        n_leading_cols=1,
        n_history_cols=0,
    )
    assert style_map == {}


def test_dataframe_generation_supports_multiple_assigned_shift_types():
    class DummySolver:
        def get_value(self, var):
            return 1 if var in {"v_d", "v_e"} else 0

        def get_objective_value(self):
            return 0

    ctx = SimpleNamespace(
        n_shift_types=2,
        shiftTypes=SimpleNamespace(
            items=[SimpleNamespace(id="D"), SimpleNamespace(id="E")],
            groups=[],
        ),
        people=SimpleNamespace(items=[SimpleNamespace(id="n1", history=None)]),
        dates=SimpleNamespace(
            items=[
                SimpleNamespace(
                    year=2025, month=1, day=1, weekday=lambda: 2, strftime=lambda fmt: "Wed" if fmt == "%a" else "1"
                )
            ],
            groups=[],
            range=SimpleNamespace(
                startDate=SimpleNamespace(year=2025, month=1), endDate=SimpleNamespace(year=2025, month=1)
            ),
        ),
        map_dp_s={(0, 0): {0, 1}},
        shifts={(0, 0, 0): "v_d", (0, 1, 0): "v_e"},
        offs={(0, 0): "v_off"},
        preferences=[],
        map_sid_s={},
        map_pid_p={},
        map_did_d={},
        solver=DummySolver(),
        solver_status="OPTIMAL",
        export=None,
    )

    df, info = exporter.get_people_versus_date_dataframe(ctx, prettify=False)
    assert df.iloc[2, 1] == "D, E"
    assert info["styles"] == {}


def test_prettify_styling_does_not_add_default_freeday_or_weekend_colors():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-03
    endDate: 2025-01-05
  groups:
    - id: WORKDAY
      members: [2025-01-03]
    - id: FREEDAY
      members: [2025-01-05]
people:
  items:
    - id: n1
      history: [D]
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
  - type: shift request
    person: n1
    date: ["2025-01-03"]
    shiftType: [OFF]
    weight: -5
"""
    styled_df, _solution, _score, _status, _cell_export_info = schedule(yaml_content, prettify=True)
    html = styled_df.to_html()
    assert "text-align: center" in html
    assert "background-color: #fefce8" not in html
    assert "#dcfce7" not in html
    assert "#dbeafe" not in html
    assert "#9ca3af" not in html
