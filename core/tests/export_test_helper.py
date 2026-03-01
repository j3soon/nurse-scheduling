"""Shared test helper for XLSX export regression tests across solver backends."""

# This file is part of Nurse Scheduling Project, see <https://github.com/j3soon/nurse-scheduling>.
#
# Copyright (C) 2023-2026 Johnson Sun
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

# This test is mostly AI generated.

import glob
import logging
import os
import sys
from io import BytesIO

# Add the project root to the Python path so imports will work when running directly
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import nurse_scheduling
import nurse_scheduling.exporter as exporter
import pandas
import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from .schedule_test_helper import CONTINUE_ON_ERROR, IGNORE_TESTS, TESTCASES_DIR


def _extract_dataframe(df_like):
    if isinstance(df_like, pandas.DataFrame):
        return df_like
    if hasattr(df_like, "data"):
        return df_like.data
    raise TypeError(f"Unsupported dataframe-like object: {type(df_like)!r}")


def _normalize_cell(value):
    return "" if value is None else str(value)


def _assert_workbook_matches_dataframe(df_like, workbook_bytes):
    expected_df = _extract_dataframe(df_like)
    wb = load_workbook(workbook_bytes)
    ws = wb.active

    assert ws.freeze_panes == "B3"
    assert ws.max_row == expected_df.shape[0]
    assert ws.max_column == expected_df.shape[1]

    for r in range(expected_df.shape[0]):
        for c in range(expected_df.shape[1]):
            expected = _normalize_cell(expected_df.iloc[r, c])
            actual = _normalize_cell(ws.cell(row=r + 1, column=c + 1).value)
            assert actual == expected, (
                f"Cell mismatch at R{r+1}C{c+1}: expected={expected!r}, actual={actual!r}"
            )


def run_export_xlsx_regression_test(solver: str, prettify: bool) -> None:
    tests = glob.glob(f"{TESTCASES_DIR}/**/*.yaml", recursive=True)
    total_tests = len(tests)
    error_count = 0

    for filepath in tests:
        base_filepath = os.path.splitext(os.path.basename(filepath))[0]
        test_dir = os.path.dirname(filepath)
        if base_filepath in IGNORE_TESTS:
            continue
        logging.info(
            "[%s][prettify=%s] Testing XLSX '%s' ...",
            solver,
            prettify,
            filepath[len(TESTCASES_DIR) + 1:],
        )

        with open(filepath, "rb") as f:
            file_content = f.read()

        # If test should fail, preserve parity with schedule regression behavior.
        if os.path.isfile(f"{test_dir}/{base_filepath}.txt"):
            with open(f"{test_dir}/{base_filepath}.txt", "r") as f:
                expected_err = f.read()
            with pytest.raises((ValidationError, ValueError)) as exc_info:
                nurse_scheduling.schedule(file_content, solver=solver, prettify=prettify)
            logging.info(f"Expected error: {expected_err.strip()}")
            logging.info(f"Actual error: {str(exc_info.value)}")
            assert expected_err.strip() in str(exc_info.value), (
                f"Expected error '{expected_err.strip()}' not found in actual error: {str(exc_info.value)}"
            )
            continue

        try:
            df, _solution, _score, _status, cell_export_info = nurse_scheduling.schedule(
                file_content,
                solver=solver,
                prettify=prettify,
            )
            if df is None:
                # Infeasible/no-solution scenarios have no table to export as XLSX.
                continue
            output = BytesIO()
            exporter.export_to_excel(df, output, cell_export_info)
            _assert_workbook_matches_dataframe(df, output)
        except ValidationError as e:
            logging.debug(f"Validation error for '{base_filepath}': {e}")
            error_count += 1
            if not CONTINUE_ON_ERROR:
                pytest.fail(f"Validation error for '{base_filepath}'")
            continue
        except Exception as e:
            logging.debug(f"Unexpected error for '{base_filepath}': {e}")
            error_count += 1
            if not CONTINUE_ON_ERROR:
                pytest.fail(f"Unexpected error for '{base_filepath}'")
            continue

    if error_count > 0:
        pytest.fail(f"Found {error_count}/{total_tests} errors during XLSX export testing")
    else:
        logging.info("All %s tests passed for solver=%s prettify=%s", total_tests, solver, prettify)
