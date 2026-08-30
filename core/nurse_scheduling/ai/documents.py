"""Bounded text extraction for untrusted AI document attachments."""

# This file is part of Nurse Scheduling Project, see <https://github.com/j3soon/nurse-scheduling>.
#
# Copyright (C) 2023-2026 Johnson Sun
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

# This code is mostly AI generated.

import json
from dataclasses import dataclass, field
from datetime import date, datetime, time
from io import BytesIO
from pathlib import PurePath
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.xml.functions import DEFUSEDXML
from pypdf import PdfReader

TEXT_DOCUMENT_EXTENSIONS = frozenset({".txt", ".md", ".csv"})
MAX_XLSX_ARCHIVE_ENTRIES = 1_000

# openpyxl uses defusedxml when it is installed before openpyxl is imported.
# Ref: https://openpyxl.readthedocs.io/en/stable/#security
if not DEFUSEDXML:
    raise RuntimeError("defusedxml is required for untrusted XLSX attachments")


class InvalidDocumentError(ValueError):
    """The uploaded bytes are not a supported, readable document."""


class DocumentLimitError(ValueError):
    """The document exceeds a configured extraction limit."""


@dataclass(frozen=True)
class DocumentExtractionLimits:
    """Limits applied while converting one attachment to prompt text."""

    max_text_chars: int
    max_pdf_pages: int
    max_xlsx_sheets: int
    max_xlsx_cells: int
    max_xlsx_uncompressed_bytes: int


@dataclass
class _BoundedText:
    """Build extracted text without silently truncating document content."""

    max_chars: int
    parts: list[str] = field(default_factory=list)
    length: int = 0

    def append(self, text: str) -> None:
        separator_length = 1 if self.parts else 0
        if self.length + separator_length + len(text) > self.max_chars:
            raise DocumentLimitError("Document extracted text is too large.")
        self.parts.append(text)
        self.length += separator_length + len(text)

    def render(self) -> str:
        return "\n".join(self.parts)


def extract_document_text(filename: str, data: bytes, limits: DocumentExtractionLimits) -> str:
    """Convert a supported document to bounded text without executing content."""
    extension = PurePath(filename).suffix.lower()
    if extension in TEXT_DOCUMENT_EXTENSIONS:
        return _extract_utf8_text(data, limits.max_text_chars)
    if extension == ".pdf":
        return _extract_pdf_text(data, limits)
    if extension == ".xlsx":
        return _extract_xlsx_text(data, limits)
    raise InvalidDocumentError("Unsupported document type.")


def _extract_utf8_text(data: bytes, max_text_chars: int) -> str:
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise InvalidDocumentError("Text document attachment must be UTF-8.") from exc
    if "\x00" in text:
        raise InvalidDocumentError("Text document attachment must be UTF-8.")
    if len(text) > max_text_chars:
        raise DocumentLimitError("Document extracted text is too large.")
    return text


def _extract_pdf_text(data: bytes, limits: DocumentExtractionLimits) -> str:
    if not data.startswith(b"%PDF-"):
        raise InvalidDocumentError("PDF content does not match its filename.")

    try:
        reader = PdfReader(BytesIO(data))
        if reader.is_encrypted:
            raise InvalidDocumentError("Encrypted PDF attachments are not supported.")
        page_count = len(reader.pages)
        if page_count > limits.max_pdf_pages:
            raise DocumentLimitError("PDF attachment has too many pages.")

        output = _BoundedText(limits.max_text_chars)
        for index, page in enumerate(reader.pages, start=1):
            # pypdf extracts embedded text but does not OCR image-only pages.
            # Ref: https://pypdf.readthedocs.io/en/stable/user/extract-text.html
            page_text = (page.extract_text() or "").strip()
            output.append(f"--- PDF page {index} ---")
            output.append(page_text or "[No extractable text. OCR was not performed.]")
        return output.render()
    except (InvalidDocumentError, DocumentLimitError):
        raise
    except Exception as exc:
        raise InvalidDocumentError("PDF attachment is invalid or unsupported.") from exc


def _extract_xlsx_text(data: bytes, limits: DocumentExtractionLimits) -> str:
    _validate_xlsx_archive(data, limits.max_xlsx_uncompressed_bytes)

    try:
        # data_only=False exposes formulas. data_only=True exposes cached results last
        # saved by a spreadsheet application and does not calculate formulas.
        # Ref: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.reader.excel.html
        formula_workbook = load_workbook(BytesIO(data), read_only=True, data_only=False, keep_links=False)
        cached_workbook = load_workbook(BytesIO(data), read_only=True, data_only=True, keep_links=False)
        try:
            if len(formula_workbook.sheetnames) > limits.max_xlsx_sheets:
                raise DocumentLimitError("XLSX attachment has too many sheets.")

            output = _BoundedText(limits.max_text_chars)
            cell_span = 0
            for sheet_name in formula_workbook.sheetnames:
                formula_sheet = formula_workbook[sheet_name]
                cached_sheet = cached_workbook[sheet_name]
                max_row = formula_sheet.max_row
                max_column = formula_sheet.max_column
                if max_row is None or max_column is None:
                    raise InvalidDocumentError("XLSX worksheet dimensions are missing.")
                cell_span += max_row * max_column
                if cell_span > limits.max_xlsx_cells:
                    raise DocumentLimitError("XLSX attachment contains too many cells.")

                output.append(f"--- XLSX sheet {json.dumps(sheet_name, ensure_ascii=False)} ---")
                populated_cells = 0
                formula_rows = formula_sheet.iter_rows(
                    min_row=1,
                    max_row=max_row,
                    min_col=1,
                    max_col=max_column,
                )
                cached_rows = cached_sheet.iter_rows(
                    min_row=1,
                    max_row=max_row,
                    min_col=1,
                    max_col=max_column,
                )
                for formula_row, cached_row in zip(formula_rows, cached_rows, strict=True):
                    for formula_cell, cached_cell in zip(formula_row, cached_row, strict=True):
                        value = formula_cell.value
                        cached_value = cached_cell.value
                        if value is None and cached_value is None:
                            continue
                        populated_cells += 1
                        if formula_cell.data_type == "f":
                            # openpyxl wraps array formulas in an object whose text holds the formula.
                            formula_value = getattr(value, "text", value)
                            cached_text = (
                                _format_cell_value(cached_value) if cached_value is not None else "unavailable"
                            )
                            output.append(
                                f"{formula_cell.coordinate}: formula {_format_cell_value(formula_value)}, "
                                f"cached value {cached_text}"
                            )
                        else:
                            output.append(f"{formula_cell.coordinate}: {_format_cell_value(value)}")
                if populated_cells == 0:
                    output.append("[No populated cells.]")
            return output.render()
        finally:
            formula_workbook.close()
            cached_workbook.close()
    except (InvalidDocumentError, DocumentLimitError):
        raise
    except Exception as exc:
        raise InvalidDocumentError("XLSX attachment is invalid or unsupported.") from exc


def _validate_xlsx_archive(data: bytes, max_uncompressed_bytes: int) -> None:
    if not data.startswith(b"PK\x03\x04"):
        raise InvalidDocumentError("XLSX content does not match its filename.")
    try:
        with ZipFile(BytesIO(data)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
                raise DocumentLimitError("XLSX attachment contains too many archive entries.")
            if any(entry.flag_bits & 0x1 for entry in entries):
                raise InvalidDocumentError("Encrypted XLSX attachments are not supported.")
            if sum(entry.file_size for entry in entries) > max_uncompressed_bytes:
                raise DocumentLimitError("XLSX attachment expands beyond the allowed size.")
            names = {entry.filename for entry in entries}
            if not {"[Content_Types].xml", "xl/workbook.xml"}.issubset(names):
                raise InvalidDocumentError("XLSX content does not match its filename.")
    except BadZipFile as exc:
        raise InvalidDocumentError("XLSX attachment is invalid or unsupported.") from exc


def _format_cell_value(value: object) -> str:
    if isinstance(value, (datetime, date, time)):
        value = value.isoformat()
    return json.dumps(value, ensure_ascii=False, default=str)
