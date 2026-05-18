"""Schedule export helpers for CSV/XLSX and rendered outputs."""

# This file is part of Nurse Scheduling Project, see <https://github.com/j3soon/nurse-scheduling>.
#
# Copyright (C) 2023-2026 Johnson Sun
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

from io import BytesIO, StringIO
from copy import copy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles.borders import Side

from .context import Context
from . import utils, models, constants


def _get_font_color_for_background(hex_color: str) -> str:
    """Return ARGB font color (black/white) for readable contrast on a hex background."""
    r = int(hex_color[1:3], 16)
    g = int(hex_color[3:5], 16)
    b = int(hex_color[5:7], 16)
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    # Match frontend threshold used in getPickerDisplay().
    return "FF000000" if luminance > 0.6 else "FFFFFFFF"


def _build_custom_export_style_info(
    ctx: Context,
    n_rows: int,
    n_cols: int,
    n_leading_rows: int,
    n_leading_cols: int,
    n_history_cols: int,
):
    """Build cell-level style overrides from ctx.export.formatting."""
    if not ctx.export or not ctx.export.formatting:
        return {}

    style_map = {}

    def set_style(
        row_idx: int,
        col_idx: int,
        background_color: str | None,
        bottom_border_color: str | None,
        right_border_color: str | None,
    ):
        if row_idx < 0 or row_idx >= n_rows or col_idx < 0 or col_idx >= n_cols:
            return
        key = (row_idx + 1, col_idx + 1)  # Store in 1-based Excel coordinates
        if key not in style_map:
            style_map[key] = {}
        if background_color:
            style_map[key]["backgroundColor"] = background_color
        if bottom_border_color:
            style_map[key]["bottomBorderColor"] = bottom_border_color
        if right_border_color:
            style_map[key]["rightBorderColor"] = right_border_color

    for rule in ctx.export.formatting:
        target_people = set()
        target_dates = set()
        target_shift_types = set()

        if rule.type in ("row", "row header", "history", "cell"):
            for target in rule.people:
                if target not in ctx.map_pid_p:
                    raise ValueError(
                        f"Invalid person identifier '{target}' in export formatting rule with type '{rule.type}'"
                    )
                target_people.update(ctx.map_pid_p[target])

        if rule.type in ("column", "column header", "cell"):
            for target in rule.dates:
                target_dates.update(utils.parse_dates(target, ctx.map_did_d, ctx.dates.range))

        if rule.type == "cell":
            for target in rule.shiftTypes:
                if target not in ctx.map_sid_s:
                    raise ValueError(
                        f"Invalid shift type identifier '{target}' in export formatting rule with type 'cell'"
                    )
                # Filter to actual shift types (exclude OFF pseudo shift).
                target_shift_types.update(s for s in ctx.map_sid_s[target] if 0 <= s < ctx.n_shift_types)

        if rule.type == "row":
            for p in target_people:
                row_idx = n_leading_rows + p
                for col_idx in range(n_cols):
                    set_style(
                        row_idx,
                        col_idx,
                        rule.backgroundColor,
                        rule.bottomBorderColor,
                        rule.rightBorderColor,
                    )

        elif rule.type == "row header":
            for p in target_people:
                row_idx = n_leading_rows + p
                set_style(row_idx, 0, rule.backgroundColor, rule.bottomBorderColor, rule.rightBorderColor)

        elif rule.type == "column":
            score_row_idx = n_leading_rows + len(ctx.people.items)
            status_row_idx = score_row_idx + 1
            for d in target_dates:
                col_idx = n_leading_cols + n_history_cols + d
                for row_idx in range(n_rows):
                    if row_idx in (score_row_idx, status_row_idx):
                        # Skip styling for score/status summary rows since they are not part of the main schedule grid and should not be affected by column styles.
                        continue
                    set_style(
                        row_idx,
                        col_idx,
                        rule.backgroundColor,
                        rule.bottomBorderColor,
                        rule.rightBorderColor,
                    )

        elif rule.type == "column header":
            for d in target_dates:
                col_idx = n_leading_cols + n_history_cols + d
                set_style(0, col_idx, rule.backgroundColor, rule.bottomBorderColor, rule.rightBorderColor)

        elif rule.type == "cell":
            for d in target_dates:
                for p in target_people:
                    if (d, p) not in ctx.map_dp_s:
                        continue
                    assigned_shift_types = [
                        s for s in ctx.map_dp_s[(d, p)] if ctx.solver.get_value(ctx.shifts[(d, s, p)]) == 1
                    ]
                    if any(s in target_shift_types for s in assigned_shift_types):
                        row_idx = n_leading_rows + p
                        col_idx = n_leading_cols + n_history_cols + d
                        set_style(
                            row_idx,
                            col_idx,
                            rule.backgroundColor,
                            rule.bottomBorderColor,
                            rule.rightBorderColor,
                        )

        elif rule.type == "history":
            for p in target_people:
                row_idx = n_leading_rows + p
                for col_idx in range(n_leading_cols, n_leading_cols + n_history_cols):
                    set_style(
                        row_idx,
                        col_idx,
                        rule.backgroundColor,
                        rule.bottomBorderColor,
                        rule.rightBorderColor,
                    )

    return style_map


def _count_extra_column_for_person(ctx: Context, p: int, count_dates, count_shift_types) -> int:
    count = 0
    for d in count_dates:
        if constants.OFF_sid in count_shift_types and ctx.solver.get_value(ctx.offs[(d, p)]) == 1:
            count += 1
            continue
        if any(
            0 <= s < ctx.n_shift_types and ctx.solver.get_value(ctx.shifts[(d, s, p)]) == 1 for s in count_shift_types
        ):
            count += 1
    return count


def _count_extra_row_for_date(ctx: Context, d: int, count_people, count_shift_types) -> int:
    count = 0
    for p in count_people:
        if constants.OFF_sid in count_shift_types and ctx.solver.get_value(ctx.offs[(d, p)]) == 1:
            count += 1
            continue
        if any(
            0 <= s < ctx.n_shift_types and ctx.solver.get_value(ctx.shifts[(d, s, p)]) == 1 for s in count_shift_types
        ):
            count += 1
    return count


def get_people_versus_date_dataframe(ctx: Context, prettify: bool = False):
    # Initialize dataframe with size including leading rows and columns
    n_leading_rows, n_leading_cols = 2, 1
    n_trailing_rows, n_trailing_cols = 2, 0

    # Dictionary to track cells with [X] markers and their weights for Excel notes
    cell_comment_info = {}

    n_history_cols = 0
    # Add history columns after the name column (only if prettify is enabled)
    if prettify:
        max_history_length = max((len(person.history) for person in ctx.people.items if person.history), default=0)
        n_history_cols = max_history_length

    extra_column_rules = ctx.export.extraColumns if prettify and ctx.export else []
    extra_row_rules = ctx.export.extraRows if prettify and ctx.export else []
    # Add extra columns and rows for prettify mode
    extra_cols = (1 + len(extra_column_rules)) if extra_column_rules else 0  # Empty separator + configured columns
    extra_rows = (1 + len(extra_row_rules)) if extra_row_rules else 0  # Empty separator + configured rows

    df = pd.DataFrame(
        "",
        index=range(n_leading_rows + len(ctx.people.items) + n_trailing_rows + extra_rows),
        columns=range(n_leading_cols + n_history_cols + len(ctx.dates.items) + n_trailing_cols + extra_cols),
        # We could cast every write to str, but object dtype is much simpler for mixed cells.
        dtype=object,
    )

    # Fill history column headers (only if prettify is enabled)
    if n_history_cols > 0:
        # - row 0 contains history position labels (H-1, H-2, etc.)
        # - row 1 contains "History" label
        for h in range(n_history_cols):
            df.iloc[0, n_leading_cols + h] = f"H-{n_history_cols - h}"
            df.iloc[1, n_leading_cols + h] = "History"

    # Fill day numbers and weekdays
    # - row 0 contains day number
    # - row 1 contains weekday
    for d, date in enumerate(ctx.dates.items):
        col_idx = n_leading_cols + n_history_cols + d
        if ctx.dates.items[0].year != ctx.dates.items[-1].year:
            df.iloc[0, col_idx] = date.strftime("%Y/%-m/%-d")
        elif ctx.dates.items[0].month != ctx.dates.items[-1].month:
            df.iloc[0, col_idx] = date.strftime("%-m/%-d")
        else:
            df.iloc[0, col_idx] = date.day
        df.iloc[1, col_idx] = date.strftime("%a")

    # Fill person descriptions and history
    # - column 0 contains person description
    # - columns 1 to n_history_cols contain history data (padded with empty strings, only if prettify)
    for p, person in enumerate(ctx.people.items):
        df.iloc[n_leading_rows + p, 0] = person.id

        # Fill history columns with proper padding (only if prettify is enabled)
        if n_history_cols > 0:
            if person.history:
                history = person.history
                # Pad with empty strings at the front if history is shorter than n_history_cols
                padded_history = [""] * max(0, n_history_cols - len(history)) + history
                for h in range(n_history_cols):
                    df.iloc[n_leading_rows + p, n_leading_cols + h] = padded_history[h]
            else:
                # Fill with empty strings if no history
                for h in range(n_history_cols):
                    df.iloc[n_leading_rows + p, n_leading_cols + h] = ""

    # Pre-filter preferences to avoid repeated filtering in the inner loop
    filtered_preferences = {}
    if prettify:
        for pref in ctx.preferences:
            if pref.type != models.SHIFT_REQUEST:
                continue
            if pref.weight == 0:
                continue
            ds = utils.parse_dates(pref.date, ctx.map_did_d, ctx.dates.range)
            ss = utils.parse_sids(pref.shiftType, ctx.map_sid_s)
            ps = utils.parse_pids(pref.person, ctx.map_pid_p)
            if len(pref.shiftType) != 1 or len(ps) != 1:
                # Skip since is not single person and single shift type style
                continue
            date_input_len = len(pref.date) if isinstance(pref.date, list) else 1
            if len(ds) != date_input_len:
                # Skip since only count for single-date style
                continue

            # Store filtered preferences by (d, p) key for quick lookup
            for d in ds:
                for p in ps:
                    if (d, p) not in filtered_preferences:
                        filtered_preferences[(d, p)] = []
                    filtered_preferences[(d, p)].append(
                        {"pref": pref, "ss": ss, "target_value": 1 if pref.weight > 0 else 0}
                    )

    # Set cell values based on solver results
    solver = ctx.solver

    for d, p in ctx.map_dp_s.keys():
        col_idx = n_leading_cols + n_history_cols + d
        assert df.iloc[n_leading_rows + p, col_idx] == ""
        cell_value = ""
        for s in ctx.map_dp_s[(d, p)]:
            if solver.get_value(ctx.shifts[(d, s, p)]) == 1:
                if cell_value != "":
                    cell_value += ", "
                cell_value += ctx.shiftTypes.items[s].id
        if prettify:
            # Only consider single-person, single-shift-type, list-of-single-date style shift request
            # Add a ` [OFF]` suffix if the person requests OFF
            # Add a ` [<shift type id>]` suffix if the person requests a specific shift type
            # Add a ` [X]` suffix if the shift request is violated
            # Use pre-filtered preferences for this (d, p) combination
            if (d, p) in filtered_preferences:
                for pref_data in filtered_preferences[(d, p)]:
                    pref = pref_data["pref"]
                    ss = pref_data["ss"]
                    target_value = pref_data["target_value"]
                    # Does not support shift type groups with mixed OFF and non-OFF shift types,
                    # which in most cases should not happen.
                    vars = [ctx.shifts[(d, s, p)] for s in ss] if constants.OFF_sid not in ss else [ctx.offs[(d, p)]]
                    if constants.OFF_sid in ss:
                        cell_value += " [OFF]"
                    else:
                        assert len(pref.shiftType) == 1
                        cell_value += f" [{pref.shiftType[0]}]"
                    if all((solver.get_value(var) != target_value) for var in vars):
                        cell_value += " [X]"
                        # Track this cell for Excel notes - store the weight
                        excel_row = n_leading_rows + p + 1  # +1 for 1-based Excel indexing
                        excel_col = n_leading_cols + n_history_cols + d + 1  # +1 for 1-based Excel indexing
                        if (excel_row, excel_col) not in cell_comment_info:
                            cell_comment_info[(excel_row, excel_col)] = []
                        cell_comment_info[(excel_row, excel_col)].append(abs(pref.weight))
        df.iloc[n_leading_rows + p, col_idx] = cell_value

    # Fill objective value
    df.iloc[n_leading_rows + len(ctx.people.items), 0] = "Score"
    df.iloc[n_leading_rows + len(ctx.people.items), n_leading_cols + n_history_cols] = solver.get_objective_value()
    # Fill solver status
    df.iloc[n_leading_rows + len(ctx.people.items) + 1, 0] = "Status"
    df.iloc[n_leading_rows + len(ctx.people.items) + 1, n_leading_cols + n_history_cols] = ctx.solver_status

    # Sanity check with offs variables
    if not prettify:
        for d, p in ctx.offs.keys():
            col_idx = n_leading_cols + n_history_cols + d
            if solver.get_value(ctx.offs[(d, p)]) == 1:
                assert df.iloc[n_leading_rows + p, col_idx] == ""
            else:
                assert df.iloc[n_leading_rows + p, col_idx] != ""

    if prettify:
        extra_col_start = n_leading_cols + n_history_cols + len(ctx.dates.items) + 1
        for rule_idx, rule in enumerate(extra_column_rules):
            col_idx = extra_col_start + rule_idx
            count_dates = utils.parse_dates(rule.countDates, ctx.map_did_d, ctx.dates.range)
            count_shift_types = utils.parse_sids(rule.countShiftTypes, ctx.map_sid_s)
            df.iloc[1, col_idx] = rule.header
            for p in range(len(ctx.people.items)):
                df.iloc[n_leading_rows + p, col_idx] = _count_extra_column_for_person(
                    ctx,
                    p,
                    count_dates,
                    count_shift_types,
                )

        extra_row_start = n_leading_rows + len(ctx.people.items) + n_trailing_rows + 1
        for rule_idx, rule in enumerate(extra_row_rules):
            row_idx = extra_row_start + rule_idx
            count_people = utils.parse_pids(rule.countPeople, ctx.map_pid_p)
            count_shift_types = utils.parse_sids(rule.countShiftTypes, ctx.map_sid_s)
            df.iloc[row_idx, 0] = rule.header
            for d in range(len(ctx.dates.items)):
                df.iloc[row_idx, n_leading_cols + n_history_cols + d] = _count_extra_row_for_date(
                    ctx,
                    d,
                    count_people,
                    count_shift_types,
                )

    # Apply default styling and borders if prettify is enabled
    if prettify:
        # Create a styler object to apply conditional formatting
        def apply_styling(df):
            # Create a style DataFrame with the same shape as the original
            style_df = pd.DataFrame("", index=df.index, columns=df.columns)

            # Apply center alignment to all cells
            for row_idx in range(len(df)):
                for col_idx in range(len(df.columns)):
                    style_df.iloc[row_idx, col_idx] = "text-align: center"

            # Apply dark red font color to cells containing violation markers "[X]"
            for row_idx in range(len(df)):
                for col_idx in range(len(df.columns)):
                    cell_value = df.iloc[row_idx, col_idx]
                    if cell_value and isinstance(cell_value, str) and "[X]" in cell_value:
                        existing_style = style_df.iloc[row_idx, col_idx]
                        if existing_style:
                            style_df.iloc[row_idx, col_idx] = f"{existing_style}; color: #C00000"
                        else:
                            style_df.iloc[row_idx, col_idx] = "color: #C00000"

            # Add borders to separate regions
            # Horizontal borders
            header_row_end = n_leading_rows - 1  # End of header region
            people_row_end = header_row_end + len(ctx.people.items)  # End of people region
            summary_row_end = people_row_end + n_trailing_rows  # End of summary region
            extra_rows_end = summary_row_end + len(extra_row_rules) + 1

            # Vertical borders
            name_col_end = n_leading_cols - 1  # End of name column
            history_col_end = name_col_end + n_history_cols  # End of history columns
            date_col_end = history_col_end + len(ctx.dates.items)  # End of date columns
            extra_columns_end = date_col_end + len(extra_column_rules) + 1

            # Apply borders to all cells, then add specific border styles
            for row_idx in range(len(df)):
                for col_idx in range(len(df.columns)):
                    base_style = style_df.iloc[row_idx, col_idx]
                    borders = []

                    # Add horizontal borders
                    if row_idx in [
                        header_row_end,
                        people_row_end,
                        summary_row_end,
                        extra_rows_end,
                    ]:
                        borders.append("border-bottom: 2px solid #374151")

                    # Add vertical borders
                    if col_idx in [
                        name_col_end,
                        history_col_end,
                        date_col_end,
                        extra_columns_end,
                    ]:
                        borders.append("border-right: 2px solid #374151")

                    # Combine base style with borders
                    if borders:
                        border_style = "; ".join(borders)
                        if base_style:
                            style_df.iloc[row_idx, col_idx] = f"{base_style}; {border_style}"
                        else:
                            style_df.iloc[row_idx, col_idx] = border_style

            return style_df

        # Apply the styling and return the styled DataFrame
        styled_df = df.style.apply(lambda x: apply_styling(df), axis=None)
        style_info = _build_custom_export_style_info(
            ctx, len(df.index), len(df.columns), n_leading_rows, n_leading_cols, n_history_cols
        )
        return styled_df, {"comments": cell_comment_info, "styles": style_info}

    style_info = _build_custom_export_style_info(
        ctx, len(df.index), len(df.columns), n_leading_rows, n_leading_cols, n_history_cols
    )
    return df, {"comments": cell_comment_info, "styles": style_info}


def export_to_excel(df, output_buffer, cell_export_info=None):
    """
    Export DataFrame to Excel with frozen panes at B3 (first two rows and first column).
    Also adds notes/comments to cells with [X] markers showing the weight of unmet single-style requests.

    Args:
        output_buffer: BytesIO buffer to write to
        cell_export_info: Dictionary containing cell comment information
    """

    # Write to a temporary BytesIO buffer first
    temp_buffer = BytesIO()
    df.to_excel(temp_buffer, index=False, header=False)
    temp_buffer.seek(0)

    # Load the workbook to apply additional formatting
    wb = load_workbook(temp_buffer)
    ws = wb.active

    # Freeze the first two rows and first column (B3 is the cell after frozen area)
    ws.freeze_panes = "B3"

    # Backward compatibility: legacy shape was {(row, col): [weights]}.
    comment_info = {}
    style_info = {}
    if isinstance(cell_export_info, dict):
        if "comments" in cell_export_info or "styles" in cell_export_info:
            comment_info = cell_export_info.get("comments") or {}
            style_info = cell_export_info.get("styles") or {}
        else:
            comment_info = cell_export_info

    # Add notes/comments to cells with [X] markers if comment_info is provided.
    if comment_info:
        from openpyxl.comments import Comment

        for (row, col), weights in comment_info.items():
            cell = ws.cell(row=row, column=col)
            # Calculate total weight and create note text
            total_weight = sum(weights)
            if len(weights) == 1:
                note_text = f"Weight of unmet single-style request: {total_weight}"
            else:
                note_text = f"Weights of unmet single-style requests: {total_weight} (individual weights: {', '.join(map(str, weights))})"

            # Create and add the comment
            comment = Comment(note_text, "Nurse Scheduling System")
            cell.comment = comment

    # Apply custom export formatting styles.
    if style_info:
        for (row, col), styles in style_info.items():
            cell = ws.cell(row=row, column=col)

            background_color = styles.get("backgroundColor")
            if background_color:
                argb = f"FF{background_color[1:].upper()}"
                cell.fill = PatternFill(fill_type="solid", start_color=argb, end_color=argb)
                updated_font = copy(cell.font)
                updated_font.color = _get_font_color_for_background(background_color)
                cell.font = updated_font

            bottom_border_color = styles.get("bottomBorderColor")
            right_border_color = styles.get("rightBorderColor")
            if bottom_border_color or right_border_color:
                existing_border = copy(cell.border)
                existing_bottom = copy(existing_border.bottom)
                existing_right = copy(existing_border.right)
                bottom_style = existing_bottom.style if existing_bottom is not None else None
                right_style = existing_right.style if existing_right is not None else None
                cell.border = Border(
                    left=existing_border.left,
                    right=(
                        Side(style=right_style or "medium", color=f"FF{right_border_color[1:].upper()}")
                        if right_border_color
                        else existing_border.right
                    ),
                    top=existing_border.top,
                    bottom=(
                        Side(style=bottom_style or "medium", color=f"FF{bottom_border_color[1:].upper()}")
                        if bottom_border_color
                        else existing_border.bottom
                    ),
                    diagonal=existing_border.diagonal,
                    diagonal_direction=existing_border.diagonal_direction,
                    outline=existing_border.outline,
                    vertical=existing_border.vertical,
                    horizontal=existing_border.horizontal,
                )

    # Save to the output buffer
    wb.save(output_buffer)
    output_buffer.seek(0)


def export_to_csv(df, output_buffer):
    """
    Export DataFrame to CSV with UTF-8 BOM for Excel compatibility.

    Args:
        output_buffer: BytesIO buffer to write to (use BytesIO for proper encoding handling)
    """
    # Write CSV to a StringIO first to get text, then encode with BOM
    temp_buffer = StringIO()
    df.to_csv(temp_buffer, index=False, header=False)
    temp_buffer.seek(0)

    # Encode with UTF-8 BOM and write to output buffer
    csv_content = temp_buffer.getvalue()
    output_buffer.write(csv_content.encode("utf-8-sig"))
    output_buffer.seek(0)
