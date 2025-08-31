import sys
import argparse
import logging
import os.path
from . import scheduler
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# TODO: Better CLI
# Ref: https://packaging.python.org/en/latest/guides/creating-command-line-tools/

def main():
    parser = argparse.ArgumentParser(description='Nurse Scheduling Tool')
    parser.add_argument('input_file_path', help='Path to the input file')
    parser.add_argument('output_path', nargs='?', help='Path to save the output file (optional)')
    parser.add_argument('--prettify', action='store_true',
                       help='Enable prettify mode for enhanced output formatting')
    parser.add_argument('-v', '--verbose', action='count', default=0,
                       help='Increase verbosity (can be used multiple times: -v, -vv, -vvv)')
    
    args = parser.parse_args()
    filepath = args.input_file_path
    output_path = args.output_path
    prettify = args.prettify
    verbose = args.verbose
    
    # Configure logging based on verbosity level
    if verbose >= 2:
        logging.basicConfig(level=logging.DEBUG, format='%(levelname)s: %(message)s')
    elif verbose == 1:
        logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
    else:
        logging.basicConfig(level=logging.WARNING, format='%(levelname)s: %(message)s')
    
    # Infer output format from file extension
    output_format = None
    if output_path:
        file_ext = os.path.splitext(output_path)[1].lower()
        if file_ext == '.xlsx':
            output_format = 'xlsx'
        elif file_ext == '.csv':
            if prettify:
                print("Error: Prettify mode is not supported for CSV files")
                sys.exit(1)
            output_format = 'csv'
        elif file_ext and file_ext not in ['.csv', '.xlsx']:
            print(f"Error: Unsupported output file extension '{file_ext}'. Supported formats: .csv, .xlsx")
            sys.exit(1)
    
    df, solution, score, status = scheduler.schedule(filepath, prettify=prettify)

    if df is None:
        print("No solution found")
        sys.exit(0)
    
    if output_path:
        # Save DataFrame in the specified format
        if output_format == 'xlsx':
            # Save to Excel with advanced formatting
            df.to_excel(output_path, index=False, header=False)
            
            # Load the workbook to apply additional formatting
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Apply center alignment to all cells
            center_alignment = Alignment(horizontal='center')
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = center_alignment
            
            # Freeze the first two rows and first column (B3 is the cell after frozen area)
            ws.freeze_panes = 'B3'
            
            # Save the formatted workbook
            wb.save(output_path)
        else:  # csv format
            # Save DataFrame to CSV with UTF-8 BOM for Excel compatibility
            df.to_csv(output_path, index=False, header=False, encoding='utf-8-sig')
        print(f"Results saved to {output_path}")
        print(f"Score: {score}")
        print(f"Status: {status}")
    else:
        print(df, solution, score, status)

if __name__ == "__main__":
    main()
